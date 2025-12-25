"""
Dışa Aktarım API Endpoint'leri
CSV, Excel ve JSON formatlarında export
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional, List
import io
import csv
import json
from datetime import datetime
import tempfile
import os

from app.core.database import get_db
from app.schemas.business import ExportRequest, ExportFormat, BusinessFilter
from app.services.business_service import BusinessService

router = APIRouter()


def sanitize_value(value):
    """Değeri export için temizle"""
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


@router.get("/download/{format}")
async def download_businesses(
    format: str,
    db: AsyncSession = Depends(get_db)
):
    """
    İşletmeleri indir (GET - kolay kullanım)
    
    Format: xlsx, csv, json
    """
    if format not in ["xlsx", "csv", "json"]:
        raise HTTPException(status_code=400, detail="Geçersiz format. xlsx, csv veya json kullanın.")
    
    service = BusinessService(db)
    filters = BusinessFilter(per_page=10000)
    businesses, total = await service.list_businesses(filters)
    
    if not businesses:
        raise HTTPException(status_code=404, detail="Dışa aktarılacak işletme bulunamadı")
    
    default_fields = [
        "name", "category", "phone", "website", "address", 
        "city", "district", "rating", "rating_count", "google_maps_url"
    ]
    
    if format == "csv":
        return await export_csv(businesses, default_fields)
    elif format == "xlsx":
        return await export_excel(businesses, default_fields)
    else:
        return await export_json(businesses, default_fields)


@router.post("/")
async def export_businesses(
    request: ExportRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    İşletmeleri dışa aktar
    
    Desteklenen formatlar:
    - **csv**: Virgülle ayrılmış değerler
    - **xlsx**: Excel dosyası
    - **json**: JSON formatı
    """
    
    service = BusinessService(db)
    
    # Filtreleri uygula veya tüm verileri al
    if request.filters:
        filters = request.filters
        filters.per_page = 10000  # Export için yüksek limit
    else:
        filters = BusinessFilter(per_page=10000)
    
    businesses, total = await service.list_businesses(filters)
    
    if not businesses:
        raise HTTPException(status_code=404, detail="Dışa aktarılacak işletme bulunamadı")
    
    # Varsayılan alanlar
    default_fields = [
        "name", "category", "phone", "website", "address", 
        "city", "district", "rating", "rating_count", "google_maps_url"
    ]
    
    fields = request.include_fields or default_fields
    
    # Format'a göre export
    if request.format == ExportFormat.CSV:
        return await export_csv(businesses, fields)
    elif request.format == ExportFormat.EXCEL:
        return await export_excel(businesses, fields)
    elif request.format == ExportFormat.JSON:
        return await export_json(businesses, fields)
    else:
        raise HTTPException(status_code=400, detail="Geçersiz format")


async def export_csv(businesses, fields: List[str]) -> StreamingResponse:
    """CSV formatında export"""
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Türkçe başlıklar
    header_map = {
        "name": "İşletme Adı",
        "category": "Kategori",
        "phone": "Telefon",
        "formatted_phone": "Telefon (Formatli)",
        "website": "Web Sitesi",
        "address": "Adres",
        "city": "Şehir",
        "district": "İlçe",
        "postal_code": "Posta Kodu",
        "rating": "Puan",
        "rating_count": "Yorum Sayısı",
        "google_maps_url": "Google Maps Linki",
        "latitude": "Enlem",
        "longitude": "Boylam"
    }
    
    # Header yaz
    headers = [header_map.get(f, f) for f in fields]
    writer.writerow(headers)
    
    # Satırları yaz
    for business in businesses:
        row = []
        for field in fields:
            value = getattr(business, field, "")
            row.append(sanitize_value(value))
        writer.writerow(row)
    
    output.seek(0)
    
    # Dosya adı
    filename = f"isletmeler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "text/csv; charset=utf-8-sig"
        }
    )


async def export_excel(businesses, fields: List[str]) -> FileResponse:
    """Excel formatında export"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="Excel desteği için openpyxl kütüphanesi gerekli"
        )
    
    # Yeni workbook oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "İşletmeler"
    
    # Türkçe başlıklar
    header_map = {
        "name": "İşletme Adı",
        "category": "Kategori",
        "phone": "Telefon",
        "formatted_phone": "Telefon (Formatli)",
        "website": "Web Sitesi",
        "address": "Adres",
        "city": "Şehir",
        "district": "İlçe",
        "postal_code": "Posta Kodu",
        "rating": "Puan",
        "rating_count": "Yorum Sayısı",
        "google_maps_url": "Google Maps Linki",
        "latitude": "Enlem",
        "longitude": "Boylam"
    }
    
    # Stiller
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header yaz
    for col, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col, value=header_map.get(field, field))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Satırları yaz
    for row_num, business in enumerate(businesses, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(business, field, "")
            cell = ws.cell(row=row_num, column=col_num, value=sanitize_value(value))
            cell.border = thin_border
            
            # Link ise hyperlink ekle
            if field in ["website", "google_maps_url"] and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
    
    # Sütun genişliklerini ayarla
    for col in range(1, len(fields) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Geçici dosyaya kaydet
    filename = f"isletmeler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)
    
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


async def export_json(businesses, fields: List[str]) -> StreamingResponse:
    """JSON formatında export"""
    
    data = []
    for business in businesses:
        item = {}
        for field in fields:
            value = getattr(business, field, None)
            if isinstance(value, datetime):
                value = value.isoformat()
            item[field] = value
        data.append(item)
    
    output = json.dumps(data, ensure_ascii=False, indent=2)
    
    filename = f"isletmeler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    return StreamingResponse(
        iter([output]),
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/fields")
async def get_export_fields():
    """
    Export edilebilir alanları listele
    """
    return {
        "fields": [
            {"key": "name", "label": "İşletme Adı", "default": True},
            {"key": "category", "label": "Kategori", "default": True},
            {"key": "phone", "label": "Telefon", "default": True},
            {"key": "formatted_phone", "label": "Telefon (Formatli)", "default": False},
            {"key": "website", "label": "Web Sitesi", "default": True},
            {"key": "address", "label": "Adres", "default": True},
            {"key": "city", "label": "Şehir", "default": True},
            {"key": "district", "label": "İlçe", "default": False},
            {"key": "postal_code", "label": "Posta Kodu", "default": False},
            {"key": "rating", "label": "Puan", "default": True},
            {"key": "rating_count", "label": "Yorum Sayısı", "default": True},
            {"key": "google_maps_url", "label": "Google Maps Linki", "default": True},
            {"key": "latitude", "label": "Enlem", "default": False},
            {"key": "longitude", "label": "Boylam", "default": False},
        ]
    }
